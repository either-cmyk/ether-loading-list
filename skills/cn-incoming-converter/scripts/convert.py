#!/usr/bin/env python3
"""
CN인사이더 출고내역(S-QED/S-QDE) → 사방넷 간편입고 등록 변환 (자립형)

- 상품코드는 스크립트 옆 product_codes.json(라벨명→상품코드)을 내장 사용.
- 별도의 '공산품 상품코드 엑셀' 업로드가 필요 없다.
- 출고내역 라벨명(H열)을 정규화하여 상품코드를 찾고, 바코드(E열)는
  출고내역 SKU(F열)에서 그대로 가져온다.
- 옵션: --codes <경로>로 외부 매핑(json 또는 공산품 xlsx C/D/F열)을 덮어쓸 수 있다.

사용법:
  python convert.py <출고내역.xlsx> [출력.xlsx] [--codes <매핑파일>]
"""

import sys
import os
import re
import json
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side
from collections import defaultdict


def norm_name(s):
    """라벨명 정규화: 모든 공백 제거 + 끝 콤마 제거"""
    s = str(s).strip()
    s = re.sub(r'\s+', '', s)
    s = s.rstrip(',')
    return s


def load_embedded_codes(script_dir):
    path = os.path.join(script_dir, 'product_codes.json')
    with open(path, encoding='utf-8') as f:
        data = json.load(f)
    return data['by_norm_name']  # {정규화라벨명: 상품코드}


def load_codes_override(override_path):
    """외부 매핑 로드: json(by_norm_name) 또는 공산품 xlsx(C상품코드/D상품명/F바코드 → 라벨명 기반)"""
    if override_path.lower().endswith('.json'):
        with open(override_path, encoding='utf-8') as f:
            d = json.load(f)
        return d.get('by_norm_name', d)
    # 공산품 xlsx: D열(상품명) 정규화 → C열(상품코드)
    wb = openpyxl.load_workbook(override_path)
    ws = wb.active
    m = {}
    for r in range(2, ws.max_row + 1):
        code = ws.cell(r, 3).value
        name = ws.cell(r, 4).value
        if code and name:
            m[norm_name(name)] = str(code).strip()
    return m


def parse_shipment(shipment_path):
    wb = openpyxl.load_workbook(shipment_path, data_only=True)
    ws = wb.active

    title = str(ws.cell(1, 1).value or '')
    m = re.search(r'(S-Q[A-Z]{2}\d+)', title)  # S-QED / S-QDE 등
    code = m.group(1) if m else 'S-QED'

    rows = []
    for r in range(3, ws.max_row + 1):
        box = ws.cell(r, 2).value
        sku = ws.cell(r, 6).value
        label = ws.cell(r, 8).value
        qk = ws.cell(r, 11).value
        ql = ws.cell(r, 12).value
        if not label or str(label).strip() == '':
            continue
        if box and str(box).strip().upper() == 'TOTAL':
            continue
        qty = ql if ql and float(ql) > 0 else qk
        qty = int(float(qty)) if qty else 0
        rows.append({
            'box': str(box).strip() if box else '',
            'sku': str(sku).strip() if sku else '',
            'label': str(label).strip(),
            'qty': qty,
        })
    return rows, code


def group_products(rows, code_map):
    products = defaultdict(lambda: {'boxes': set(), 'qty': 0, 'label': '', 'code': '', 'barcode': ''})
    unmatched = []
    for row in rows:
        key = row['sku'] or row['label']  # 바코드 우선 그룹핑
        code = code_map.get(norm_name(row['label']), '')
        if not code and row['label'] not in [u['label'] for u in unmatched]:
            unmatched.append({'sku': row['sku'], 'label': row['label']})
        p = products[key]
        p['boxes'].add(row['box'])
        p['qty'] += row['qty']
        p['label'] = row['label']
        p['code'] = code
        p['barcode'] = row['sku']
    return sorted(products.values(), key=lambda x: x['label']), unmatched


def format_box_range(boxes):
    nums, prefix = [], ''
    for b in boxes:
        parts = b.rsplit('-', 1)
        if len(parts) == 2:
            prefix = parts[0]
            try:
                nums.append(int(parts[1]))
            except ValueError:
                continue
    if not nums:
        return ', '.join(sorted(boxes))
    nums.sort()
    ranges, i = [], 0
    while i < len(nums):
        start = end = nums[i]
        while i + 1 < len(nums) and nums[i + 1] == end + 1:
            i += 1
            end = nums[i]
        ranges.append(f"{prefix}-{start:03d}" if start == end
                      else f"{prefix}-{start:03d}~{prefix}-{end:03d}")
        i += 1
    return ', '.join(ranges)


def create_output(products, output_path):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = '엑셀입고'
    headers = ['박스넘버', '출고상품코드', '상품명', '수량', '바코드', '유통기한', '로케이션', '입고메모']
    border = Border(left=Side(style='thin'), right=Side(style='thin'),
                    top=Side(style='thin'), bottom=Side(style='thin'))
    for c, h in enumerate(headers, 1):
        cell = ws.cell(1, c, h)
        cell.font = Font(bold=True, size=10, name='맑은 고딕')
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = border
    for i, p in enumerate(products, 2):
        data = [format_box_range(p['boxes']), p['code'], p['label'], p['qty'], p['barcode'], None, None, None]
        for c, v in enumerate(data, 1):
            cell = ws.cell(i, c, v)
            cell.font = Font(size=10, name='맑은 고딕')
            cell.border = border
    widths = {'A': 30, 'B': 15, 'C': 55, 'D': 8, 'E': 18, 'F': 12, 'G': 12, 'H': 12}
    for col, w in widths.items():
        ws.column_dimensions[col].width = w
    wb.save(output_path)


def main():
    args = [a for a in sys.argv[1:]]
    override = None
    if '--codes' in args:
        idx = args.index('--codes')
        override = args[idx + 1]
        del args[idx:idx + 2]
    if not args:
        print("사용법: python convert.py <출고내역.xlsx> [출력.xlsx] [--codes <매핑>]")
        sys.exit(1)
    shipment_path = args[0]
    output_path = args[1] if len(args) > 1 else None

    script_dir = os.path.dirname(os.path.abspath(__file__))
    code_map = load_codes_override(override) if override else load_embedded_codes(script_dir)
    print(f"상품코드 매핑: {len(code_map)}건 ({'외부' if override else '내장'})")

    rows, code = parse_shipment(shipment_path)
    print(f"출고내역: {len(rows)}행 (코드 {code})")

    products, unmatched = group_products(rows, code_map)
    if unmatched:
        print("\n⚠️ 라벨명 매칭 실패(상품코드 빈칸):")
        for u in unmatched:
            print(f"  - [{u['sku']}] {u['label']}")

    if not output_path:
        biz = '이더컴퍼니'
        base = os.path.basename(shipment_path)
        for p in base.split('_'):
            if any(x in p for x in ('컴퍼니', '인테크', '플로', '코퍼레이션')):
                biz = p.replace('주식회사', '').replace('주식회사 ', '').strip()
                break
        output_path = f"{code}_{biz}_간편입고등록.xlsx"

    create_output(products, output_path)

    total = sum(p['qty'] for p in products)
    independent = sum((r['qty']) for r in rows)
    print(f"\n결과: {len(products)}개 상품, 총 수량 {total}")
    for p in products:
        print(f"  {format_box_range(p['boxes'])} | {p['code'] or '(미매칭)'} | {p['label']} | {p['qty']}")
    if total == independent:
        print(f"\n✅ 내부 합계 검증 통과 (그룹합 {total} = 행합 {independent})")
    else:
        print(f"\n⚠️ 합계 불일치! 그룹합 {total} ≠ 행합 {independent}")
    if unmatched:
        print(f"⚠️ 미매칭 {len(unmatched)}건 — 상품코드 수동 확인 필요")
    print(f"✅ 저장: {output_path}")


if __name__ == '__main__':
    main()
