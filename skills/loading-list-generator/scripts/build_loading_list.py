#!/usr/bin/env python3
"""
쿠팡 팔레트 적재리스트 빌더 (v1.1.0)

v1.1.0 변경점:
- 헤더 이름 기반 열 인덱싱 (이미지 컬럼 변형 자동 대응)
- N-G 트레이 / 혼합 트레이 지원
- 자동 혼적 감지 + 사용자 명시 화이트리스트 합집합
- PDF 행 높이 자동 조정 (행 수에 따라 9/8.5/7.5/7mm)
- 거래명세서 보정 (ADJUSTMENTS)

사용법:
  build_loading_list.py --config config.json --output-dir <마운트 폴더>

config.json 스키마는 SKILL.md 참고.
"""
import argparse
import json
from collections import OrderedDict, defaultdict
from datetime import datetime
from pathlib import Path

import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.page import PageMargins

from reportlab.lib.pagesizes import A4
from reportlab.lib import colors
from reportlab.lib.units import mm
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, PageBreak
from reportlab.lib.styles import ParagraphStyle


def detect_columns(ws):
    """헤더(row 2) 이름 기반 컬럼 인덱스 감지 (이미지 컬럼 변형 대응)"""
    headers = [str(ws.cell(2, c).value or "") for c in range(1, ws.max_column + 1)]
    return {
        "tray": headers.index("트레이 번호") + 1,
        "box": headers.index("상자 번호") + 1,
        "sku": headers.index("SKU") + 1,
        "ship": headers.index("출고수량") + 1,
        "set": headers.index("세트수량") + 1,
    }


def detect_mixed_boxes(src_xlsx, tray_to_pallet):
    """같은 (tray, box_code) 다중 SKU = 자동 혼적 감지"""
    wb = openpyxl.load_workbook(src_xlsx, data_only=True)
    ws = wb.active
    cols = detect_columns(ws)
    box_skus = defaultdict(set)
    for r in range(3, ws.max_row + 1):
        tray = ws.cell(r, cols["tray"]).value
        box = ws.cell(r, cols["box"]).value
        sku = ws.cell(r, cols["sku"]).value
        if not box or not sku or not tray:
            continue
        if str(tray).strip() in ("TOTAL", "트레이 수", "상자 수"):
            continue
        t = str(tray).strip()
        if t not in tray_to_pallet:
            continue
        box_skus[(t, str(box).strip())].add(str(sku).strip())
    return {b for (t, b), skus in box_skus.items() if len(skus) > 1}


def parse_sqed(src_xlsx, tray_to_pallet, true_mixed):
    """S-QED 엑셀 파싱 → pallet_boxes[p] = [(box_code, OrderedDict(bc->qty)), ...]"""
    wb = openpyxl.load_workbook(src_xlsx, data_only=True)
    ws = wb.active
    cols = detect_columns(ws)
    pallet_boxes = {p: [] for p in set(tray_to_pallet.values())}
    mixed_tracker = {}

    for r in range(3, ws.max_row + 1):
        tray = ws.cell(r, cols["tray"]).value
        box = ws.cell(r, cols["box"]).value
        bc = ws.cell(r, cols["sku"]).value
        ship = ws.cell(r, cols["ship"]).value
        setq = ws.cell(r, cols["set"]).value
        if not box or not bc or not tray:
            continue
        if str(tray).strip() in ("TOTAL", "트레이 수", "상자 수"):
            continue
        t = str(tray).strip()
        b = str(box).strip()
        bcs = str(bc).strip()
        if t not in tray_to_pallet:
            continue
        qty = int(setq) if setq else int(ship) if ship else 0
        if qty == 0:
            continue
        p = tray_to_pallet[t]
        if b in true_mixed:
            key = (p, b)
            if key in mixed_tracker:
                pallet_boxes[p][mixed_tracker[key]][1][bcs] = (
                    pallet_boxes[p][mixed_tracker[key]][1].get(bcs, 0) + qty
                )
            else:
                od = OrderedDict()
                od[bcs] = qty
                pallet_boxes[p].append((b, od))
                mixed_tracker[key] = len(pallet_boxes[p]) - 1
        else:
            od = OrderedDict()
            od[bcs] = qty
            pallet_boxes[p].append((b, od))

    return pallet_boxes


def build_rows(pallet_no, pallet_boxes, true_mixed, bc_to_product, adjustments, declared_box_count):
    """파레트별 합본 행 생성"""
    boxes = pallet_boxes.get(pallet_no, [])
    phys = len(boxes)
    declared = declared_box_count.get(str(pallet_no), declared_box_count.get(pallet_no, phys))

    totals = OrderedDict()
    mixed_marks = {}
    for idx, (bc_box, skus) in enumerate(boxes, 1):
        is_mixed = bc_box in true_mixed
        for bc, q in skus.items():
            totals[bc] = totals.get(bc, 0) + q
            if is_mixed:
                mixed_marks.setdefault(bc, []).append(f"{declared}-{idx}")

    for adj in adjustments:
        if adj["pallet"] == pallet_no and adj["bc"] in totals:
            totals[adj["bc"]] += adj["delta"]

    def sort_key(bc):
        sku = bc_to_product.get(bc, ["99999999", ""])[0]
        return int(sku)

    rows = []
    for bc in sorted(totals.keys(), key=sort_key):
        sku, name = bc_to_product.get(bc, ["", bc])
        marks = mixed_marks.get(bc, [])
        seen = set()
        uniq = []
        for m in marks:
            if m not in seen:
                seen.add(m)
                uniq.append(m)
        rows.append((sku, name, ", ".join(uniq) if uniq else "", totals[bc]))

    return declared, rows


def get_row_height_mm(n_rows):
    """행 수에 따라 PDF 행 높이 자동 조정 (A4 1페이지 fit)"""
    if n_rows <= 12: return 9
    if n_rows <= 16: return 8.5
    if n_rows <= 20: return 7.5
    return 7


# ----------- XLSX -----------
def build_xlsx(out_path, cfg, pallet_data):
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    thin = Side(style="thin", color="000000")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    gray_fill = PatternFill("solid", fgColor="F2F2F2")

    for p in sorted(pallet_data.keys()):
        pc, rows_data = pallet_data[p]
        ws_new = wb.create_sheet(f"팔레트 {cfg['total_pallets']}-{p}")
        ws_new.page_setup.orientation = "portrait"
        ws_new.page_setup.paperSize = ws_new.PAPERSIZE_A4
        ws_new.page_setup.fitToWidth = 1
        ws_new.page_setup.fitToHeight = 1
        ws_new.sheet_properties.pageSetUpPr.fitToPage = True
        ws_new.page_margins = PageMargins(left=0.4, right=0.4, top=0.5, bottom=0.4)
        for i, w in enumerate([6, 14, 28, 24, 14, 11, 18], 1):
            ws_new.column_dimensions[get_column_letter(i)].width = w

        ws_new.merge_cells("A1:G1")
        ws_new["A1"] = "쿠팡 팔레트 적재리스트 (각 팔레트 부착 필수)"
        ws_new["A1"].font = Font(name="맑은 고딕", size=18, bold=True)
        ws_new["A1"].alignment = Alignment(horizontal="center", vertical="center")
        ws_new.row_dimensions[1].height = 30
        ws_new.merge_cells("A2:G2")
        ws_new["A2"] = "※ 팔레트의 높이는 1,700mm를 초과할 수 없습니다 ※"
        ws_new["A2"].font = Font(name="맑은 고딕", size=12, bold=True)
        ws_new["A2"].alignment = Alignment(horizontal="center", vertical="center")
        ws_new.row_dimensions[2].height = 22
        ws_new["A4"] = "1) 업체 정보"
        ws_new["A4"].font = Font(name="맑은 고딕", size=11, bold=True)

        def info(row, k, v, bold=False):
            ws_new.merge_cells(start_row=row, start_column=1, end_row=row, end_column=2)
            ws_new.cell(row, 1, k).font = Font(name="맑은 고딕", size=10)
            ws_new.cell(row, 1).alignment = Alignment(horizontal="center", vertical="center")
            ws_new.cell(row, 1).fill = gray_fill
            ws_new.cell(row, 1).border = border
            ws_new.cell(row, 2).border = border
            ws_new.merge_cells(start_row=row, start_column=3, end_row=row, end_column=7)
            ws_new.cell(row, 3, v).font = Font(name="맑은 고딕", size=10, bold=bold)
            ws_new.cell(row, 3).alignment = Alignment(horizontal="left", vertical="center", indent=1)
            for c in range(3, 8):
                ws_new.cell(row, c).border = border
            ws_new.row_dimensions[row].height = 22

        info(5, "업체명", cfg["company_name"])
        info(6, "업체코드", cfg["company_code"])
        ws_new["A8"] = "2) 입고 예약 정보"
        ws_new["A8"].font = Font(name="맑은 고딕", size=11, bold=True)
        for i, (k, v, b) in enumerate([
            ("요청 ID", cfg["order_id"], False),
            ("물류센터", cfg["center"], False),
            ("물류센터 도착예정일", cfg["arrival_date"], False),
            ("팔레트 번호", f"{cfg['total_pallets']}-{p}", True),
            ("총 박스", f"{pc}", False),
        ]):
            info(9 + i, k, v, b)
        ws_new["A15"] = "3) 상품 정보"
        ws_new["A15"].font = Font(name="맑은 고딕", size=11, bold=True)
        ws_new.merge_cells("C16:D16")
        for c, v in enumerate(
            ["No.", "SKU ID", "물류입고용 상품명 / 옵션명", None, "박스 번호", "상품 수량", "소비기한/제조일자"], 1
        ):
            if v is None:
                continue
            cell = ws_new.cell(16, c, v)
            cell.font = Font(name="맑은 고딕", size=10, bold=True)
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.fill = gray_fill
            cell.border = border
        ws_new.cell(16, 4).border = border
        ws_new.row_dimensions[16].height = 26
        for i, row in enumerate(rows_data):
            r = 17 + i
            ws_new.cell(r, 1, i + 1)
            ws_new.cell(r, 2, row[0])
            ws_new.merge_cells(start_row=r, start_column=3, end_row=r, end_column=4)
            ws_new.cell(r, 3, row[1])
            ws_new.cell(r, 5, row[2])
            ws_new.cell(r, 6, row[3])
            ws_new.cell(r, 7, "-")
            for c in range(1, 8):
                cell = ws_new.cell(r, c)
                cell.font = Font(name="맑은 고딕", size=9)
                cell.alignment = Alignment(
                    horizontal="center" if c != 3 else "left",
                    vertical="center",
                    wrap_text=True,
                    indent=1 if c == 3 else 0,
                )
                cell.border = border
            ws_new.row_dimensions[r].height = 22
        ws_new.print_area = f"A1:G{16 + len(rows_data)}"
    wb.save(out_path)


# ----------- PDF -----------
def build_pdf(out_path, cfg, pallet_data, font_dir):
    pdfmetrics.registerFont(TTFont("Nanum", str(Path(font_dir) / "NotoSansKR-Regular.ttf")))
    pdfmetrics.registerFont(TTFont("NanumB", str(Path(font_dir) / "NotoSansKR-Bold.ttf")))
    doc = SimpleDocTemplate(
        str(out_path), pagesize=A4,
        leftMargin=12 * mm, rightMargin=12 * mm,
        topMargin=12 * mm, bottomMargin=10 * mm,
    )
    ts_t = ParagraphStyle("t", fontName="NanumB", fontSize=18, alignment=1, leading=22)
    ts_s = ParagraphStyle("s", fontName="NanumB", fontSize=12, alignment=1, leading=16, spaceAfter=8)
    ts_sec = ParagraphStyle("sec", fontName="NanumB", fontSize=11, leading=14, spaceBefore=10, spaceAfter=4)
    ts_c = ParagraphStyle("c", fontName="Nanum", fontSize=9, alignment=1, leading=11)
    ts_cb = ParagraphStyle("cb", fontName="NanumB", fontSize=10, alignment=1, leading=12)
    ts_cl = ParagraphStyle("cl", fontName="Nanum", fontSize=9, alignment=0, leading=11, leftIndent=4)
    ts_clb = ParagraphStyle("clb", fontName="NanumB", fontSize=10, alignment=0, leading=12, leftIndent=4)
    ts_n = ParagraphStyle("n", fontName="Nanum", fontSize=8, alignment=0, leading=10, leftIndent=4)

    story = []
    pallet_keys = sorted(pallet_data.keys())
    for p in pallet_keys:
        pc, rows = pallet_data[p]
        story.append(Paragraph("쿠팡 팔레트 적재리스트 (각 팔레트 부착 필수)", ts_t))
        story.append(Paragraph("※ 팔레트의 높이는 1,700mm를 초과할 수 없습니다 ※", ts_s))
        story.append(Paragraph("1) 업체 정보", ts_sec))
        t1 = Table([
            [Paragraph("업체명", ts_cb), Paragraph(cfg["company_name"], ts_cl)],
            [Paragraph("업체코드", ts_cb), Paragraph(cfg["company_code"], ts_cl)],
        ], colWidths=[45 * mm, 141 * mm], rowHeights=[8 * mm] * 2)
        t1.setStyle(TableStyle([
            ("GRID", (0, 0), (-1, -1), 0.5, colors.black),
            ("BACKGROUND", (0, 0), (0, -1), colors.HexColor("#F2F2F2")),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ]))
        story.append(t1)
        story.append(Paragraph("2) 입고 예약 정보", ts_sec))
        t2 = Table([
            [Paragraph("요청 ID", ts_cb), Paragraph(cfg["order_id"], ts_cl)],
            [Paragraph("물류센터", ts_cb), Paragraph(cfg["center"], ts_cl)],
            [Paragraph("물류센터 도착예정일", ts_cb), Paragraph(cfg["arrival_date"], ts_cl)],
            [Paragraph("팔레트 번호", ts_cb), Paragraph(f"<b>{cfg['total_pallets']}-{p}</b>", ts_clb)],
            [Paragraph("총 박스", ts_cb), Paragraph(f"{pc}", ts_cl)],
        ], colWidths=[45 * mm, 141 * mm], rowHeights=[8 * mm] * 5)
        t2.setStyle(TableStyle([
            ("GRID", (0, 0), (-1, -1), 0.5, colors.black),
            ("BACKGROUND", (0, 0), (0, -1), colors.HexColor("#F2F2F2")),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ]))
        story.append(t2)
        story.append(Paragraph("3) 상품 정보", ts_sec))
        tdata = [[Paragraph(h, ts_cb) for h in ["No.", "SKU ID", "물류입고용 상품명 / 옵션명", "박스 번호", "상품 수량", "소비기한/제조일자"]]]
        for i, row in enumerate(rows, 1):
            tdata.append([
                Paragraph(str(i), ts_c),
                Paragraph(str(row[0]), ts_c),
                Paragraph(str(row[1]), ts_n),
                Paragraph(str(row[2]) if row[2] else "", ts_c),
                Paragraph(str(row[3]), ts_c),
                Paragraph("-", ts_c),
            ])
        rh = get_row_height_mm(len(rows))
        pt = Table(
            tdata,
            colWidths=[12 * mm, 24 * mm, 76 * mm, 26 * mm, 20 * mm, 28 * mm],
            rowHeights=[10 * mm] + [rh * mm] * len(rows),
        )
        pt.setStyle(TableStyle([
            ("GRID", (0, 0), (-1, -1), 0.5, colors.black),
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#F2F2F2")),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ]))
        story.append(pt)
        if p != pallet_keys[-1]:
            story.append(PageBreak())
    doc.build(story)


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--config", required=True)
    ap.add_argument("--output-dir", required=True)
    ap.add_argument("--font-dir", default=str(Path(__file__).resolve().parent.parent / "assets"))
    args = ap.parse_args()

    with open(args.config, encoding="utf-8") as f:
        cfg = json.load(f)

    # 자동 혼적 감지 + 사용자 명시 합집합
    auto_mixed = detect_mixed_boxes(cfg["src_xlsx"], cfg["tray_to_pallet"])
    user_mixed = set(cfg.get("true_mixed", []))
    true_mixed = auto_mixed | user_mixed
    print(f"Auto-detected mixed: {sorted(auto_mixed)}")
    if user_mixed:
        print(f"User-specified mixed: {sorted(user_mixed)}")

    pallet_boxes = parse_sqed(cfg["src_xlsx"], cfg["tray_to_pallet"], true_mixed)

    pallet_data = {}
    grand_total = 0
    for p in sorted(set(cfg["tray_to_pallet"].values())):
        declared, rows = build_rows(
            p, pallet_boxes, true_mixed,
            cfg["bc_to_product"],
            cfg.get("adjustments", []),
            cfg.get("declared_box_count", {}),
        )
        pallet_data[p] = (declared, rows)
        grand_total += sum(r[3] for r in rows)
        tray = [k for k, v in cfg["tray_to_pallet"].items() if v == p][0]
        print(f"\nPallet {cfg['total_pallets']}-{p} [{tray}] ({declared} boxes, 총 {sum(r[3] for r in rows)}):")
        for sku, name, bn, q in rows:
            bnstr = f"[{bn}]" if bn else ""
            print(f"  {sku} {name[:50]:<52} {bnstr:>14} qty:{q}")

    # 거래명세서 검증
    expected_total = sum(cfg.get("delivery_total", {}).values())
    if expected_total:
        match = "✓" if grand_total == expected_total else "✗"
        print(f"\nGrand total: {grand_total} {match} (거래명세서 {expected_total})")
    else:
        print(f"\nGrand total: {grand_total}")

    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    out_dir = Path(args.output_dir)
    out_dir.mkdir(parents=True, exist_ok=True)
    xlsx_path = out_dir / f"적재리스트_{cfg['order_id']}_{ts}.xlsx"
    pdf_path = out_dir / f"적재리스트_{cfg['order_id']}_{ts}.pdf"
    build_xlsx(xlsx_path, cfg, pallet_data)
    build_pdf(pdf_path, cfg, pallet_data, args.font_dir)
    print(f"\nSaved XLSX: {xlsx_path}")
    print(f"Saved PDF:  {pdf_path}")


if __name__ == "__main__":
    main()
