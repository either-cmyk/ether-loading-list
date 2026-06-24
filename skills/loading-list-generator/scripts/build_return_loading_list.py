#!/usr/bin/env python3
"""쿠팡 회송분(반품 재작업) 적재리스트 생성 (v1.2.0 신규)

표준 입고분과 다른 점:
- 입력 엑셀이 멀티시트(시트 1개 = 파레트 1개, 예: "10-01" = 1번 파레트).
- 각 시트의 박스수량 컬럼(E열) **병합셀**이 박스 1개를 의미한다.
  (병합 그룹 안의 여러 행 = 같은 박스에 든 혼합 품목)
- 상품 식별은 "바코드끝6자리"이며, 동봉문서(거래명세서) 바코드 끝 6자리로 매칭한다.

출력은 표준과 동일: SKU/옵션당 1줄 합본 + 혼적박스만 콤마 박스번호({박스수}-{순번}).
PDF/XLSX 렌더링은 build_loading_list.py의 build_pdf/build_xlsx를 그대로 재사용한다.

config(JSON) 구조:
{
  "order_id": "134329099",
  "company_name": "주식회사 이더컴퍼니",
  "company_code": "A01139144",
  "center": "동탄1(DON1)",
  "arrival_date": "2026-06-26",
  "src_xlsx": "/path/회송엑셀.xlsx",
  "bc6_to_product": { "050254": ["59835158", "이더커머스 ... 그레이 235-240"], ... },
  "code_fix": { "055156": "050156", ... },          # (선택) 바코드 오타 보정
  "sheet_exclude": ["전체 품목합계"],                # (선택) 제외 시트
  "delivery_total": { ... }                          # (선택) 동봉 발주합 검증용
}
"""

import argparse
import json
import re
from collections import defaultdict, OrderedDict
from datetime import datetime
from pathlib import Path

import openpyxl

from build_loading_list import build_xlsx
from reportlab.lib.pagesizes import A4
from reportlab.lib.units import mm
from reportlab.lib import colors
from reportlab.platypus import SimpleDocTemplate, Paragraph, Table, TableStyle, PageBreak
from reportlab.lib.styles import ParagraphStyle
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont


def col_index(ws, names):
    """헤더(row 3) 이름으로 컬럼 인덱스 찾기. 회송 양식은 보통 row3 헤더."""
    for hr in (3, 2, 1):
        headers = [str(ws.cell(hr, c).value or "").strip() for c in range(1, ws.max_column + 1)]
        idx = {}
        for key, cands in names.items():
            for cand in cands:
                if cand in headers:
                    idx[key] = headers.index(cand) + 1
                    break
        if "code" in idx and "qty" in idx:
            return idx, hr
    raise ValueError("헤더(바코드끝6자리/총수량)를 찾지 못함")


def merge_blocks(ws, box_col):
    """box_col(박스수량) 병합 범위 → row -> (min_row,max_row). 단일행은 (r,r)."""
    m = {}
    for rng in ws.merged_cells.ranges:
        if rng.min_col <= box_col <= rng.max_col:
            for r in range(rng.min_row, rng.max_row + 1):
                m[r] = (rng.min_row, rng.max_row)
    return m


def parse_return_sheet(ws, idx, hr, code_fix):
    """시트 1개 → (box_count, boxes[ [(code,qty), ...], ... ])  (병합셀=박스)"""
    box_col = idx.get("box", 5)
    mb = merge_blocks(ws, box_col)
    boxes, seen = [], {}
    for r in range(hr + 1, ws.max_row + 1):
        code = ws.cell(r, idx["code"]).value
        if code is None or str(code).strip() in ("합계", "계", "소계"):
            continue
        qty = ws.cell(r, idx["qty"]).value
        if not isinstance(qty, (int, float)):
            continue
        code = str(code).strip().zfill(6)
        code = code_fix.get(code, code)
        grp = mb.get(r, (r, r))
        if grp not in seen:
            seen[grp] = len(boxes)
            boxes.append([])
        boxes[seen[grp]].append((code, int(qty)))
    return len(boxes), boxes


def build_pallet_data(cfg):
    wb = openpyxl.load_workbook(cfg["src_xlsx"], data_only=True)
    exclude = set(cfg.get("sheet_exclude", ["전체 품목합계"]))
    code_fix = cfg.get("code_fix", {})
    bc6 = cfg["bc6_to_product"]
    sheets = [s for s in wb.sheetnames if s not in exclude]

    pallet_data = OrderedDict()
    unmatched = []
    for seq, sn in enumerate(sheets, 1):
        ws = wb[sn]
        idx, hr = col_index(ws, {
            "code": ["바코드끝6자리", "바코드", "SKU"],
            "qty": ["총수량", "출고수량", "수량"],
            "box": ["박스수량", "박스"],
        })
        box_count, boxes = parse_return_sheet(ws, idx, hr, code_fix)
        # SKU/옵션당 1줄 합본 + 혼적박스 콤마 박스번호
        agg = OrderedDict()  # sku -> {name, dno, qty, mix[]}
        for bseq, items in enumerate(boxes, 1):
            is_mixed = len(items) > 1
            for code, qty in items:
                if code in bc6:
                    sku, name = bc6[code][0], bc6[code][1]
                    dno = bc6[code][2] if len(bc6[code]) > 2 else 99999
                else:
                    sku, name, dno = code, f"(미매칭) {code}", 99999
                    unmatched.append((sn, code, qty))
                a = agg.setdefault(sku, {"name": name, "dno": dno, "qty": 0, "mix": []})
                a["qty"] += qty
                if is_mixed:
                    a["mix"].append(bseq)
        rows = []
        for sku in sorted(agg, key=lambda s: (agg[s]["dno"], s)):
            a = agg[sku]
            boxno = ", ".join(f"{box_count}-{b}" for b in sorted(set(a["mix"])))
            rows.append((sku, a["name"], boxno, a["qty"]))
        pallet_data[seq] = (box_count, rows)
    return pallet_data, sheets, unmatched




def build_pdf_return(out_path, cfg, pallet_data, font_dir):
    """회송 전용 PDF: 박스번호 콤마 리스트가 길어 행높이를 자동(가변)으로 둔다."""
    pdfmetrics.registerFont(TTFont("Nanum", str(Path(font_dir) / "NotoSansKR-Regular.ttf")))
    pdfmetrics.registerFont(TTFont("NanumB", str(Path(font_dir) / "NotoSansKR-Bold.ttf")))
    doc = SimpleDocTemplate(str(out_path), pagesize=A4,
        leftMargin=10*mm, rightMargin=10*mm, topMargin=10*mm, bottomMargin=9*mm)
    t_t = ParagraphStyle("t", fontName="NanumB", fontSize=15, alignment=1, leading=18)
    t_s = ParagraphStyle("s", fontName="Nanum", fontSize=9, alignment=1, leading=12, spaceAfter=6)
    t_sec = ParagraphStyle("sec", fontName="NanumB", fontSize=9.5, leading=12, spaceBefore=6, spaceAfter=3)
    t_c = ParagraphStyle("c", fontName="Nanum", fontSize=8, alignment=1, leading=9.5)
    t_cb = ParagraphStyle("cb", fontName="NanumB", fontSize=8.5, alignment=1, leading=10)
    t_cl = ParagraphStyle("cl", fontName="Nanum", fontSize=8.5, alignment=0, leading=10, leftIndent=4)
    t_n = ParagraphStyle("n", fontName="Nanum", fontSize=7.3, alignment=0, leading=8.6, leftIndent=3)
    t_bx = ParagraphStyle("bx", fontName="Nanum", fontSize=7.3, alignment=1, leading=8.6)
    story = []
    keys = sorted(pallet_data.keys())
    for p in keys:
        pc, rows = pallet_data[p]
        story.append(Paragraph("쿠팡 팔레트 적재리스트 (각 팔레트 부착 필수)", t_t))
        story.append(Paragraph("※ 팔레트의 높이는 1,700mm를 초과할 수 없습니다 ※", t_s))
        story.append(Paragraph("1) 업체 정보", t_sec))
        t1 = Table([[Paragraph("업체명", t_cb), Paragraph(cfg["company_name"], t_cl)],
                    [Paragraph("업체코드", t_cb), Paragraph(cfg["company_code"], t_cl)]],
                   colWidths=[40*mm, 150*mm])
        t1.setStyle(TableStyle([("GRID",(0,0),(-1,-1),0.5,colors.black),
            ("BACKGROUND",(0,0),(0,-1),colors.HexColor("#F2F2F2")),("VALIGN",(0,0),(-1,-1),"MIDDLE"),
            ("TOPPADDING",(0,0),(-1,-1),2),("BOTTOMPADDING",(0,0),(-1,-1),2)]))
        story.append(t1)
        story.append(Paragraph("2) 입고 예약 정보", t_sec))
        t2 = Table([[Paragraph("요청 ID", t_cb), Paragraph(cfg["order_id"], t_cl)],
                    [Paragraph("물류센터", t_cb), Paragraph(cfg["center"], t_cl)],
                    [Paragraph("물류센터 도착예정일", t_cb), Paragraph(cfg.get("arrival_date",""), t_cl)],
                    [Paragraph("팔레트 번호", t_cb), Paragraph(f"<b>{cfg['total_pallets']}-{p}</b>", t_cl)],
                    [Paragraph("총 박스", t_cb), Paragraph(f"{pc}", t_cl)]],
                   colWidths=[40*mm, 150*mm])
        t2.setStyle(TableStyle([("GRID",(0,0),(-1,-1),0.5,colors.black),
            ("BACKGROUND",(0,0),(0,-1),colors.HexColor("#F2F2F2")),("VALIGN",(0,0),(-1,-1),"MIDDLE"),
            ("TOPPADDING",(0,0),(-1,-1),2),("BOTTOMPADDING",(0,0),(-1,-1),2)]))
        story.append(t2)
        story.append(Paragraph("3) 상품 정보", t_sec))
        tdata = [[Paragraph(h, t_cb) for h in ["No.","SKU ID","물류입고용 상품명 / 옵션명","박스 번호","수량","소비기한/제조"]]]
        for i, row in enumerate(rows, 1):
            tdata.append([Paragraph(str(i), t_c), Paragraph(str(row[0]), t_c),
                Paragraph(str(row[1]), t_n), Paragraph(str(row[2]) if row[2] else "", t_bx),
                Paragraph(str(row[3]), t_c), Paragraph("-", t_c)])
        pt = Table(tdata, colWidths=[8*mm, 21*mm, 64*mm, 76*mm, 13*mm, 8*mm], repeatRows=1)
        pt.setStyle(TableStyle([("GRID",(0,0),(-1,-1),0.5,colors.black),
            ("BACKGROUND",(0,0),(-1,0),colors.HexColor("#F2F2F2")),("VALIGN",(0,0),(-1,-1),"MIDDLE"),
            ("TOPPADDING",(0,0),(-1,-1),1.2),("BOTTOMPADDING",(0,0),(-1,-1),1.2)]))
        story.append(pt)
        if p != keys[-1]:
            story.append(PageBreak())
    doc.build(story)


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--config", required=True)
    ap.add_argument("--output-dir", required=True)
    ap.add_argument("--font-dir", default=str(Path(__file__).resolve().parent.parent / "assets"))
    args = ap.parse_args()

    with open(args.config, encoding="utf-8") as f:
        cfg = json.load(f)

    pallet_data, sheets, unmatched = build_pallet_data(cfg)
    cfg["total_pallets"] = len(sheets)
    # build_pdf/build_xlsx 호환 키
    cfg.setdefault("arrival_date", cfg.get("delivery_date", ""))

    grand = 0
    for p, (bc, rows) in pallet_data.items():
        s = sum(r[3] for r in rows)
        grand += s
        print(f"Pallet {cfg['total_pallets']}-{p} ({bc} boxes, 총 {s}, {len(rows)}행)")
    print(f"Grand total: {grand}")
    exp = sum(cfg.get("delivery_total", {}).values())
    if exp:
        print(f"동봉 발주합 {exp} {'✓' if grand <= exp else '✗(초과!)'}")
    if unmatched:
        print(f"⚠ 미매칭 {len(unmatched)}건: {unmatched[:10]}")

    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    out_dir = Path(args.output_dir)
    out_dir.mkdir(parents=True, exist_ok=True)
    xlsx_path = out_dir / f"적재리스트_회송_{cfg['order_id']}_{ts}.xlsx"
    pdf_path = out_dir / f"적재리스트_회송_{cfg['order_id']}_{ts}.pdf"
    build_xlsx(xlsx_path, cfg, pallet_data)
    build_pdf_return(pdf_path, cfg, pallet_data, args.font_dir)
    print(f"Saved XLSX: {xlsx_path}")
    print(f"Saved PDF:  {pdf_path}")


if __name__ == "__main__":
    main()
